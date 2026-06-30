"""
sync_xlsx.py
============
Keeps the human-facing Excel mirrors in sync with the machine source-of-truth
JSON files. The pytest suite always reads the JSON; the .xlsx exists so students
can view/edit the metadata and master cases the same way they did in Levels 1-2.

Uses openpyxl only (no pandas) so it installs cleanly on any Python 3.11+.

Usage:
  python tools/sync_xlsx.py to-xlsx data/parameter_metadata.json data/parameter_metadata.xlsx
  python tools/sync_xlsx.py to-json data/parameter_metadata.xlsx data/parameter_metadata.json
"""

import json
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook


def _flatten(value):
    """Excel cells can't hold lists/dicts — store them as JSON text."""
    if isinstance(value, (list, dict)):
        return json.dumps(value)
    return value


def _unflatten(value):
    if isinstance(value, str) and value[:1] in "[{":
        try:
            return json.loads(value)
        except json.JSONDecodeError:
            return value
    return value


def to_xlsx(json_path: str, xlsx_path: str) -> None:
    rows = json.loads(Path(json_path).read_text(encoding="utf-8"))
    if not rows:
        raise SystemExit("no rows to write")
    headers = list(rows[0].keys())
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    ws.append(headers)
    for row in rows:
        ws.append([_flatten(row.get(h)) for h in headers])
    wb.save(xlsx_path)
    print(f"wrote {len(rows)} rows -> {xlsx_path}")


def to_json(xlsx_path: str, json_path: str) -> None:
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    headers = list(next(it))
    rows = [{h: _unflatten(v) for h, v in zip(headers, values)} for values in it]
    Path(json_path).write_text(json.dumps(rows, indent=2), encoding="utf-8")
    print(f"wrote {len(rows)} rows -> {json_path}")


if __name__ == "__main__":
    if len(sys.argv) != 4 or sys.argv[1] not in {"to-xlsx", "to-json"}:
        raise SystemExit(__doc__)
    direction, src, dst = sys.argv[1], sys.argv[2], sys.argv[3]
    (to_xlsx if direction == "to-xlsx" else to_json)(src, dst)
